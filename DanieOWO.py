# Danie van den Berg
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import Font
from xls2xlsx import XLS2XLSX
import os

try:
    x2x = XLS2XLSX("C:/Users/User/Desktop/mypcauto/kragmyn.xls")
    x2x.to_xlsx("C:/Users/User/Desktop/mypcauto/kragmyn.xlsx")

    os.remove("C:/Users/User/Desktop/mypcauto/kragmyn.xls")
except:
    pass

thick_border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

grayfill = PatternFill(start_color='BFBFBF',
                       end_color='BFBFBF',
                       fill_type='solid')


try:
    wb2 = load_workbook('C:/Users/User/Desktop/mypcauto/kragmyn.xlsx')

    boolmine = True
except:
    boolmine = False
    print("failed 2")


wb = load_workbook('C:/Users/User/Desktop/mypcauto/kraghuise.xlsx')

if boolmine:
    ws2 = wb2.active
    maxrowb = ws2.max_row

ws = wb.active

maxrowi = ws.max_row


for i in range(1, maxrowi+1):  # we loop through each row
    if boolmine:

        if i == maxrowi:
            for k in range(1, maxrowb+1):

                if str(ws2['A' + str(k)].value).upper() == "METER #":
                    ws.cell(row=i+k, column=2).border = thick_border
                    ws['B' + str(i+k)].fill = grayfill
                    ws['B' + str(i+k)].font = Font(bold=True)

                if str(ws2['B' + str(k)].value).upper() == "TOKEN #":
                    ws.cell(row=i+k, column=3).border = thick_border
                    ws['C' + str(i+k)].fill = grayfill
                    ws['C' + str(i+k)].font = Font(bold=True)

                if str(ws2['D' + str(k)].value).upper() == "VALUEACT" or str(ws2['D' + str(k)].value).upper() == "VALUEACTUAL":
                    ws.cell(row=i+k, column=4).border = thick_border
                    ws['D' + str(i+k)].fill = grayfill
                    ws['D' + str(i+k)].font = Font(bold=True)

                if str(ws2['C' + str(k)].value).upper() == "RAND" or str(ws2['C' + str(k)].value).upper() == "PRICE":
                    ws.cell(row=i+k, column=5).border = thick_border
                    ws['E' + str(i+k)].fill = grayfill
                    ws['E' + str(i+k)].font = Font(bold=True)

                    ws["F" + str(i+k)].value = "House"
                    ws.cell(row=i+k, column=6).border = thick_border
                    ws['F' + str(i+k)].fill = grayfill
                    ws['F' + str(i+k)].font = Font(bold=True)

                    ws['G' + str(i+k)].value = "Sold To"
                    ws.cell(row=i+k, column=7).border = thick_border
                    ws['G' + str(i+k)].fill = grayfill
                    ws['G' + str(i+k)].font = Font(bold=True)

                    ws['H' + str(i+k)].value = "Date"
                    ws.cell(row=i+k, column=8).border = thick_border
                    ws['H' + str(i+k)].fill = grayfill
                    ws['H' + str(i+k)].font = Font(bold=True)

                    ws['I' + str(i+k)].value = "TVL REC"
                    ws.cell(row=i+k, column=9).border = thick_border
                    ws['I' + str(i+k)].fill = grayfill
                    ws['I' + str(i+k)].font = Font(bold=True)

                    ws['J' + str(i+k)].value = "Payment"
                    ws.cell(row=i+k, column=10).border = thick_border
                    ws['J' + str(i+k)].fill = grayfill
                    ws['J' + str(i+k)].font = Font(bold=True)

                if not ws2['A' + str(k)].value == None:
                    ws['B' + str(i+k)].value = ws2['A' + str(k)].value

                if not ws2['B' + str(k)].value == None:
                    ws['C' + str(i+k)].value = ws2['B' + str(k)].value

                if not ws2['D' + str(k)].value == None:
                    ws['D' + str(i+k)].value = ws2['D' + str(k)].value

                if not ws2['C' + str(k)].value == None:
                    ws['E' + str(i+k)].value = ws2['C' + str(k)].value

                if ws['B' + str(i+k)].value != "METER #" and ws['B' + str(i+k)].value != None:
                    # print(ws["B" + str(i+k)].value, i, k)
                    ws.cell(row=i+k, column=2).border = thin_border
                    ws.cell(row=i+k, column=3).border = thin_border
                    ws.cell(row=i+k, column=4).border = thin_border
                    ws.cell(row=i+k, column=5).border = thin_border
                    ws.cell(row=i+k, column=6).border = thin_border
                    ws.cell(row=i+k, column=7).border = thin_border
                    ws.cell(row=i+k, column=8).border = thin_border
                    ws.cell(row=i+k, column=9).border = thin_border
                    ws.cell(row=i+k, column=10).border = thin_border

    if str(ws['E'+str(i)].value).upper() == "PRICE" or str(ws['E'+str(i)].value).upper() == "RAND":

        ws['F' + str(i)].value = "House"
        ws.cell(row=i, column=6).border = thick_border
        ws['F'+str(i)].fill = grayfill
        ws['F' + str(i)].font = Font(bold=True)

        ws['G' + str(i)].value = "Sold To"
        ws.cell(row=i, column=7).border = thick_border
        ws['G'+str(i)].fill = grayfill
        ws['G' + str(i)].font = Font(bold=True)

        ws['H' + str(i)].value = "Date"
        ws.cell(row=i, column=8).border = thick_border
        ws['H'+str(i)].fill = grayfill
        ws['H' + str(i)].font = Font(bold=True)

        ws['I' + str(i)].value = "TVL REC"
        ws.cell(row=i, column=9).border = thick_border
        ws['I'+str(i)].fill = grayfill
        ws['I' + str(i)].font = Font(bold=True)

        ws['J' + str(i)].value = "Payment"
        ws.cell(row=i, column=10).border = thick_border
        ws['J'+str(i)].fill = grayfill
        ws['J' + str(i)].font = Font(bold=True)
    else:
        ws.cell(row=i, column=6).border = thin_border
        ws.cell(row=i, column=7).border = thin_border
        ws.cell(row=i, column=8).border = thin_border
        ws.cell(row=i, column=9).border = thin_border
        ws.cell(row=i, column=10).border = thin_border


maxrownew = ws.max_row
for i in range(1, maxrownew+1):
    if ws["B" + str(i)].value == None:
        ws.delete_rows(i)

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:

        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width
wb.save('C:/Users/User/Desktop/kragkoeponne.xlsx')


try:
    os.remove("C:/Users/User/Desktop/mypcauto/kragmyn.xlsx")
except:
    pass

try:
    os.remove("C:/Users/User/Desktop/mypcauto/kraghuise.xlsx")
except:
    pass
